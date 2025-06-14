import json
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Case, When
from django.http import JsonResponse
from django.db.models import Max
from patient.models import Patient, OtherReference, SocialMediaReference
from account.models import User
from address_place.models import Country
from store.models import Store, MedicineStoreTransactionHistory
from django.shortcuts import render, redirect
from my_order.models import NormalInvoiceTracker, EstimateInvoiceTracker, InvoiceTracker
from .models import PatientMedicineBillHead, PatientMedicineBillDetail, PatientMedicineUnregisteredBillHead, \
    PatientMedicineUnregisteredBillDetail, PatientEstimateMedicineBillHead, PatientEstimateMedicineBillDetail
from store.models import MedicineStore
from django.db.models.functions import Coalesce
from decimal import Decimal
from django.db.models import DecimalField
from django.db.models import Sum, F, FloatField, ExpressionWrapper
import requests

from django.http import HttpResponse
from openpyxl import Workbook
from .models import PatientMedicineBillHead
from openpyxl.utils import get_column_letter

from datetime import datetime, time
from django.utils import timezone


# Create your views here.
def normal_generate_invoice_number():
    current_year = datetime.now().year
    obj, _ = NormalInvoiceTracker.objects.get_or_create(year=current_year)
    invoice_number = obj.get_next_invoice_number()
    return invoice_number


def generate_invoice_number():
    current_year = datetime.now().year
    obj, _ = InvoiceTracker.objects.get_or_create(year=current_year)
    invoice_number = obj.get_next_invoice_number()
    return invoice_number


def estimate_generate_invoice_number():
    current_year = datetime.now().year
    obj, _ = EstimateInvoiceTracker.objects.get_or_create(year=current_year)
    invoice_number = obj.get_next_invoice_number()
    return invoice_number


def patient_list(request):
    patient = Patient.objects.all().order_by('-id')
    context = {
        'patient': patient
    }
    return render(request, 'patient_list.html', context)


def patient_detail(request, id):
    if request.method == 'POST':
        form = request.POST
        username = form.get('patient_name')
        care_of = form.get('care_of')
        mobile = form.get('mobile')
        patient_age = form.get('patient_age')
        sex = form.get('sex')
        house_flat = form.get('house_flat')
        street = form.get('street')
        city = form.get('city')
        district = form.get('district')
        pincode = form.get('pincode')
        country = form.get('country')
        state = form.get('state')
        reference_by_patient = form.get('patient_search_id')
        reference_by_other = form.get('reference_by_other')
        social_media = form.get('social_media')
        email = mobile + '@yopmail.com'
        phone = form.get('phone')

        tobacco = form.get('tobacco')
        alcohol = form.get('alcohol')
        smoking = form.get('smoking')
        if tobacco is not None:
            tobacco = 1
        else:
            tobacco = 0

        if alcohol is not None:
            alcohol = 1
        else:
            alcohol = 0

        if smoking is not None:
            smoking = 1
        else:
            smoking = 0

        status = 'failed'
        msg = 'Patient Registration failed.'
        try:
            user = Patient.objects.get(id=id)
            user_obj = User.objects.filter(id=user.user.id).update(username=username.title(),
                                                                   email=email,
                                                                   password='12345',
                                                                   mobile=mobile,
                                                                   phone=phone,
                                                                   care_of=care_of.title(),
                                                                   sex=sex,
                                                                   age=patient_age,
                                                                   house_flat=house_flat,
                                                                   street_colony=street,
                                                                   city=city,
                                                                   district=district,
                                                                   pin=pincode,
                                                                   state_id=state,
                                                                   country_id=country,
                                                                   )
            if user_obj:
                if social_media:
                    social_media = social_media
                else:
                    social_media = None
                if reference_by_other:
                    reference_by_other = reference_by_other
                else:
                    reference_by_other = None
                if reference_by_patient:
                    reference_by_patient = reference_by_patient
                else:
                    reference_by_patient = None
                Patient.objects.filter(id=id).update(social_media_id=social_media,
                                                     other_reference_id=reference_by_other,
                                                     reference_by_patient_id=reference_by_patient,
                                                     tobacco=tobacco,
                                                     alcohol=alcohol,
                                                     smoking=smoking,
                                                     )
                status = 'success'
                msg = 'Patient updated successfully.'

        except Exception as e:
            status = status
            msg = str(e)

        context = {
            'status': status,
            'msg': msg,
        }
        return JsonResponse(context)
    else:
        patient = Patient.objects.get(id=id)
        try:
            country = Country.objects.annotate(is_patient_country=Case(When(id=patient.user.country.id,
                                                                            then=0), default=1, )).order_by(
                'is_patient_country', 'name')
        except:
            country = Country.objects.all()

        if patient.social_media:
            reference_social_media_id = patient.social_media.id
        else:
            reference_social_media_id = None
        social_media = SocialMediaReference.objects.annotate(
            is_social_media=Case(
                When(id=reference_social_media_id, then=0),
                default=1, )
        ).order_by('is_social_media', 'title')
        if patient.reference_by_patient:
            reference_by_patient_id = patient.reference_by_patient.id
        else:
            reference_by_patient_id = None
        reference_by_other = OtherReference.objects.annotate(
            is_reference_by_other=Case(
                When(id=reference_by_patient_id, then=0),
                default=1,
            )
        ).order_by('is_reference_by_other', 'name')
        context = {
            'country': country,
            'social_media': social_media,
            'reference_by_other': reference_by_other,
            'id': id,
            'patient': patient
        }
        return render(request, 'patient_detail.html', context)


def search_patient(request):
    if 'term' in request.GET:
        search_value = request.GET.get('term')
        if search_value:
            search_terms = search_value.split()
            for term in search_terms:
                query = Q(user__username__icontains=term) | Q(patient_code__icontains=term)
                qs = Patient.objects.filter(query)
                data_list = []
                for i in qs:
                    data_dict = {}
                    data_dict['id'] = i.id
                    data_dict['name'] = i.user.username
                    data_dict['code'] = i.patient_code
                    data_list.append(data_dict)
                context = {
                    'data_list': data_list
                }
                return JsonResponse(context, safe=False)
            return JsonResponse([], safe=False)


def add_other_reference(request):
    if request.method == 'POST':
        form = request.POST
        reference_name = form.get('reference_name')
        reference_mobile = form.get('reference_mobile')
        reference_address = form.get('reference_address')
        patient_obj = OtherReference.objects.create(name=reference_name,
                                                    mobile=reference_mobile,
                                                    address=reference_address, )
        patient_dict = {
            'id': patient_obj.id,
            'name': patient_obj.name,
            'mobile': patient_obj.mobile,
        }
        if patient_obj:
            patient_obj = patient_dict
            status = 'success'
            msg = 'Reference registered successfully.'
        else:
            patient_obj = {}
            status = 'failed'
            msg = 'Reference registered failed.'

        context = {
            'status': status,
            'msg': msg,
            'patient_obj': patient_obj,
        }
        return JsonResponse(context)


def get_patient(request):
    if request.method == 'POST':
        form = request.POST
        search_data = form.get('search_data')
        status = 'failed'
        msg = 'Patient data not found?'
        patient_id = 0
        try:
            user_obj = Patient.objects.filter(user__mobile=search_data)
            if user_obj:
                patient_id = user_obj[0].id
                status = 'success'
                msg = 'Data found successfully.'
        except Exception as e:
            status = status
            msg = str(e)

        context = {
            'status': status,
            'msg': msg,
            'patient_id': patient_id,
        }
        return JsonResponse(context)
    return render(request, 'customer_bill/create_customer_bill.html')


def send_sms(mobile, message):
    url = "http://msg.msgclub.net/rest/services/sendSMS/sendGroupSms"
    params = {
        'AUTH_KEY': '3380567192fd2e6d18f63985aace',
        'message': message,
        'senderId': 'MRCARC',
        'routeId': 1,
        'mobileNos': mobile,
        'smsContentType': 'english'
    }

    response = requests.get(url, params=params)
    if response.status_code == 200:
        print("SMS sent successfully!")
    else:
        print("Failed to send SMS", response.text)


def create_patient_bill(request, patient_id=0):
    if request.method == 'POST':
        form = request.POST
        first_name = form.get('name')
        mobile = form.get('mobile')
        mobile = mobile[-10:]
        email = form.get('email')
        patient_age = form.get('age')
        sex = form.get('gender')
        tobacco = form.get('tobacco')
        alcohol = form.get('alcohol')
        smoking = form.get('smoking')
        patient_code = datetime.now().strftime("%Y%d%H%M%S")
        status = 'failed'
        msg = 'Patient Registration failed.'
        # try:
        #     user_obj = User.objects.create_user(username=mobile,
        #                                         first_name=first_name,
        #                                         email=email,
        #                                         password='12345',
        #                                         mobile=mobile,
        #                                         sex=sex,
        #                                         age=patient_age,
        #                                         )
        #     if user_obj:
        #         patient_id = user_obj.id
        #         patient_obj = Patient.objects.create(user_id=patient_id,
        #                                              patient_code=patient_code,
        #                                              )
        #         if patient_obj:
        #             patient_id = patient_obj.id
        #             status = 'success'
        #             msg = 'Patient Registration successfully.'
        #
        # except Exception as e:
        #     status = status
        #     msg = str(e)

        context = {
            'status': status,
            'msg': msg,
            'patient_id': patient_id,
        }
        return JsonResponse(context)
    else:
        country = Country.objects.all()
        social_media = SocialMediaReference.objects.all()
        reference_by_other = OtherReference.objects.all()
        context = {
            'country': country,
            'social_media': social_media,
            'reference_by_other': reference_by_other,
        }
        return render(request, 'customer_bill/create_customer_bill.html', context)


def create_patient_bill_detail(request, patient_id):
    patient = Patient.objects.get(id=patient_id)
    context = {
        'patient_id': patient_id,
        'patient': patient,
    }
    return render(request, 'customer_bill/create_customer_detail.html', context)


def patient_generate_bill(request, order_type, patient_id):
    patient = Patient.objects.get(id=patient_id)

    cash_online_amount = PatientMedicineBillHead.objects.filter(patient_id=patient_id).aggregate(
        total=Sum(
            Coalesce(F('cash'), 0) +
            Coalesce(F('online'), 0) +
            Coalesce(F('extra_cash_amount'), 0) +
            Coalesce(F('extra_online_amount'), 0),
            output_field=DecimalField()
        )
    )['total'] or 0

    total_pay_amount = PatientMedicineBillHead.objects.filter(patient_id=patient_id).aggregate(
        total=Sum('pay_amount')
    )['total'] or 0

    old_credit_sum = total_pay_amount - cash_online_amount

    context = {
        'patient_id': patient_id,
        'patient': patient,
        'order_type': order_type,
        'old_credit_sum': old_credit_sum,
    }
    if order_type == 1:
        return render(request, 'customer_bill/create_customer_bill_instate.html', context)
    elif order_type == 2:
        return render(request, 'customer_bill/create_customer_bill_other_state.html', context)
    elif order_type == 3:
        return render(request, 'customer_bill/create_customer_bill_of_supply.html', context)


@login_required(login_url='/account/user_login/')
def create_bill(request, order_type, patient_id):
    if request.method == 'POST':
        form = request.POST
        user_id = request.session['user_id']
        try:
            store = Store.objects.get(user_id=user_id)
            store_id = store.id
        except:
            store_id = 0

        status = 'failed'
        msg = 'Bill Creation failed.'

        medicines = json.loads(request.POST.get('medicines'))
        invoice_number = normal_generate_invoice_number()

        subtotal = float(form.get('sub_total'))
        flat_discount = float(form.get('flat_discount'))
        discount = int(form.get('total_discount'))
        shipping_packing = float(form.get('shipping_packing'))

        discount_amount = subtotal * discount / 100
        after_dis_amount = subtotal - discount_amount + shipping_packing
        after_dis_amount = after_dis_amount - flat_discount

        cash = float(form.get('cash'))
        online = float(form.get('online'))

        sgst = form.get('sgst')
        cgst = form.get('cgst')

        obj = PatientMedicineBillHead.objects.create(patient_id=patient_id,
                                                     store_id=store_id,
                                                     sgst=sgst,
                                                     cgst=cgst,
                                                     subtotal=subtotal,
                                                     flat_discount=flat_discount,
                                                     discount_amount=discount_amount,
                                                     cash=cash,
                                                     online=online,
                                                     shipping=shipping_packing,
                                                     discount=discount,
                                                     pay_amount=after_dis_amount,
                                                     current=after_dis_amount,
                                                     status=1,
                                                     order_type=order_type,
                                                     )
        if obj:
            mobile = obj.patient.user.mobile
            created_at = obj.created_at
            head_id = obj.id
            for medicine_data in medicines:
                medicine_id = medicine_data['medicine_id']
                batch_no = medicine_data['batch_no']
                record_qty = int(medicine_data['record_qty'])
                sell_qty = int(medicine_data['sell_qty'])
                discount = int(medicine_data['discount'])
                mrp = float(medicine_data['mrp'])
                sale_rate = float(medicine_data['sale_rate'])
                try:
                    hsn = medicine_data['hsn']
                except:
                    hsn = ''

                try:
                    gst = int(medicine_data['gst'])
                except:
                    gst = 0

                try:
                    taxable_amount = float(medicine_data['taxable_amount'])
                except:
                    taxable_amount = 0

                try:
                    tax = float(medicine_data['tax'])
                except:
                    tax = 0

                amount = float(medicine_data['amount'])

                obj = PatientMedicineBillDetail.objects.create(head_id=head_id,
                                                               medicine_id=medicine_id,
                                                               record_qty=record_qty,
                                                               sell_qty=sell_qty,
                                                               mrp=mrp,
                                                               discount=discount,
                                                               sale_rate=sale_rate,
                                                               hsn=hsn,
                                                               gst=gst,
                                                               taxable_amount=taxable_amount,
                                                               tax=tax,
                                                               amount=amount,
                                                               batch_no=batch_no,
                                                               )

                if obj:
                    if sell_qty <= record_qty:
                        remaining_qty = int(record_qty) - int(sell_qty)
                        MedicineStore.objects.filter(to_store_id=store_id, medicine_id=medicine_id,
                                                     batch_no=batch_no).update(
                            qty=remaining_qty)
                        PatientMedicineBillDetail.objects.filter(id=obj.id).update(record_qty=remaining_qty)
            PatientMedicineBillHead.objects.filter(id=head_id).update(status=1, invoice_number=invoice_number, )

            message = f'Thank you.We have received a payment of Rs Cash {cash}, online {online},old pending {100}, total pending  {after_dis_amount}... Against Bill No. {invoice_number} .. on Date at time {created_at}. Thanks for visiting Mrc Ayurveda'
            # message = f'Thank you.We have received a payment of Rs {after_dis_amount} Against Bill no.{invoice_number}, on Date at time. Thanks for visiting Mrc Ayurveda'

            try:
                send_sms(mobile, message)
                send_sms(9267678888, message)
            except:
                pass

            status = 'success'
            msg = 'Bill creation Successfully.'

        context = {
            'status': status,
            'msg': msg,
        }
        return JsonResponse(context)

    else:
        user_id = request.session['user_id']
        try:
            store = Store.objects.get(user_id=user_id)
            store_id = store.id
        except:
            store_id = 0

        context = {
            'id': id,
            'order_type': order_type,
            'store_id': store_id,
        }
        if order_type == 1:
            return render(request, 'normal_bill/order_tax_invoice_in_state.html', context)
        elif order_type == 2:
            return render(request, 'normal_bill/order_tax_invoice_other_state.html', context)
        elif order_type == 3:
            return render(request, 'normal_bill/oder_bill_of_supply.html', context)
        else:
            return render(request, 'normal_bill/oder_bill_of_supply.html', context)


@login_required(login_url='/account/user_login/')
def unregistered_create_bill(request, order_type, id):
    if request.method == 'POST':
        form = request.POST
        user_id = request.session['user_id']
        try:
            store = Store.objects.get(user_id=user_id)
            store_id = store.id
        except:
            store_id = 0

        status = 'failed'
        msg = 'Bill Creation failed.'

        medicines = json.loads(request.POST.get('medicines'))
        invoice_number = generate_invoice_number()
        patient_id = form.get('patient_id')

        subtotal = float(form.get('sub_total'))
        flat_discount = float(form.get('flat_discount'))
        discount = int(form.get('total_discount'))
        shipping_packing = float(form.get('shipping_packing'))

        discount_amount = subtotal * discount / 100
        after_dis_amount = subtotal - discount_amount + shipping_packing
        after_dis_amount = after_dis_amount - flat_discount

        cash = float(form.get('cash'))
        online = float(form.get('online'))

        sgst = form.get('sgst')
        cgst = form.get('cgst')
        state_code = form.get('state_code')
        is_already = PatientMedicineUnregisteredBillHead.objects.filter(head_id=id)
        if is_already:
            head_id = is_already[0].id
            PatientMedicineUnregisteredBillDetail.objects.filter(head_id=head_id).delete()
            PatientMedicineUnregisteredBillHead.objects.filter(head_id=id).delete()
            PatientMedicineBillHead.objects.filter(id=id).update(final_bill_status=0)

        obj = PatientMedicineUnregisteredBillHead.objects.create(head_id=id,
                                                                 invoice_number=invoice_number,
                                                                 patient_id=patient_id,
                                                                 store_id=store_id,
                                                                 sgst=sgst,
                                                                 cgst=cgst,
                                                                 subtotal=subtotal,
                                                                 flat_discount=flat_discount,
                                                                 discount_amount=discount_amount,
                                                                 cash=cash,
                                                                 online=online,
                                                                 shipping=shipping_packing,
                                                                 discount=discount,
                                                                 pay_amount=after_dis_amount,
                                                                 current=after_dis_amount,
                                                                 status=1,
                                                                 order_type=order_type,
                                                                 state_code=state_code,
                                                                 )
        if obj:
            head_id = obj.id
            for medicine_data in medicines:
                medicine_id = medicine_data['medicine_id']
                batch_no = medicine_data['batch_no']
                record_qty = int(medicine_data['record_qty'])
                sell_qty = int(medicine_data['sell_qty'])
                discount = int(medicine_data['discount'])
                mrp = float(medicine_data['mrp'])
                sale_rate = float(medicine_data['sale_rate'])
                try:
                    hsn = medicine_data['hsn']
                except:
                    hsn = ''

                try:
                    gst = int(medicine_data['gst'])
                except:
                    gst = 0

                try:
                    taxable_amount = float(medicine_data['taxable_amount'])
                except:
                    taxable_amount = 0

                try:
                    tax = float(medicine_data['tax'])
                except:
                    tax = 0

                amount = float(medicine_data['amount'])

                obj = PatientMedicineUnregisteredBillDetail.objects.create(head_id=head_id,
                                                                           medicine_id=medicine_id,
                                                                           record_qty=record_qty,
                                                                           sell_qty=sell_qty,
                                                                           mrp=mrp,
                                                                           discount=discount,
                                                                           sale_rate=sale_rate,
                                                                           hsn=hsn,
                                                                           gst=gst,
                                                                           taxable_amount=taxable_amount,
                                                                           tax=tax,
                                                                           amount=amount,
                                                                           batch_no=batch_no,
                                                                           )
                PatientMedicineBillHead.objects.filter(id=id).update(final_bill_status=1)
                if obj:
                    status = 'success'
                    msg = 'Bill creation Successfully.'

            PatientMedicineUnregisteredBillHead.objects.filter(id=head_id).update(invoice_number=invoice_number)

        context = {
            'status': status,
            'msg': msg,
        }
        return JsonResponse(context)


@login_required(login_url='/account/user_login/')
def update_medicine_order_bill(request, order_type, id):
    if request.method == 'POST':
        form = request.POST
        user_id = request.session['user_id']
        try:
            store = Store.objects.get(user_id=user_id)
            store_id = store.id
        except:
            store_id = 0

        status = 'failed'
        msg = 'Bill Creation failed.'
        medicines = json.loads(request.POST.get('medicines'))

        subtotal = float(form.get('sub_total'))
        flat_discount = float(form.get('flat_discount'))
        discount1 = int(form.get('total_discount'))
        shipping_packing = float(form.get('shipping_packing'))

        discount_amount = subtotal * discount1 / 100
        after_dis_amount = subtotal - discount_amount + shipping_packing
        after_dis_amount = after_dis_amount - flat_discount

        cash = float(form.get('cash'))
        online = float(form.get('online'))

        sgst = form.get('sgst')
        cgst = form.get('cgst')
        state_code = form.get('state_code')

        head_id = id

        medicine_ids = []
        for medicine_data in medicines:
            medicine_id = medicine_data['medicine_id']
            batch_no = medicine_data['batch_no']
            medicine_ids.append(medicine_id)
            record_qty = int(medicine_data['record_qty'])
            sell_qty = int(medicine_data['sell_qty'])
            discount = int(medicine_data['discount'])
            mrp = float(medicine_data['mrp'])
            sale_rate = float(medicine_data['sale_rate'])
            try:
                hsn = medicine_data['hsn']
            except:
                hsn = 0

            try:
                gst = int(medicine_data['gst'])
            except:
                gst = 0

            try:
                taxable_amount = float(medicine_data['taxable_amount'])
            except:
                taxable_amount = 0

            try:
                tax = float(medicine_data['tax'])
            except:
                tax = 0
            amount = float(medicine_data['amount'])

            already_obj = PatientMedicineBillDetail.objects.filter(head_id=head_id, medicine_id=medicine_id,
                                                                   batch_no=batch_no)

            if already_obj:
                pre_sell_qty = already_obj[0].sell_qty
                update_record_qty = int(record_qty + pre_sell_qty)
                MedicineStore.objects.filter(to_store_id=store_id, medicine_id=medicine_id, batch_no=batch_no).update(
                    qty=update_record_qty)

                store_record = MedicineStore.objects.filter(to_store_id=store_id, medicine_id=medicine_id,
                                                            batch_no=batch_no)

                if store_record[0].qty >= sell_qty:
                    record_qty = int(store_record[0].qty) - sell_qty
                    PatientMedicineBillDetail.objects.filter(head_id=head_id, medicine_id=medicine_id,
                                                             batch_no=batch_no).update(record_qty=record_qty,
                                                                                       sell_qty=sell_qty,
                                                                                       mrp=mrp,
                                                                                       discount=discount,
                                                                                       sale_rate=sale_rate,
                                                                                       hsn=hsn,
                                                                                       gst=gst,
                                                                                       taxable_amount=taxable_amount,
                                                                                       tax=tax,
                                                                                       amount=amount,
                                                                                       batch_no=batch_no,
                                                                                       )

                    MedicineStore.objects.filter(to_store_id=store_id, medicine_id=medicine_id,
                                                 batch_no=batch_no).update(
                        qty=record_qty)
            else:
                store_record = MedicineStore.objects.filter(to_store_id=store_id, medicine_id=medicine_id,
                                                            batch_no=batch_no)
                if store_record[0].qty >= sell_qty:
                    record_qty = int(store_record[0].qty) - sell_qty
                    PatientMedicineBillDetail.objects.create(head_id=head_id,
                                                             medicine_id=medicine_id,
                                                             record_qty=record_qty,
                                                             sell_qty=sell_qty,
                                                             mrp=mrp,
                                                             discount=discount,
                                                             sale_rate=sale_rate,
                                                             hsn=hsn,
                                                             gst=gst,
                                                             taxable_amount=taxable_amount,
                                                             tax=tax,
                                                             amount=amount,
                                                             batch_no=batch_no,
                                                             )
                    MedicineStore.objects.filter(to_store_id=store_id, medicine_id=medicine_id,
                                                 batch_no=batch_no).update(qty=record_qty)

        obj = PatientMedicineBillHead.objects.filter(id=id).update(store_id=store_id,
                                                                   sgst=sgst,
                                                                   cgst=cgst,
                                                                   subtotal=subtotal,
                                                                   cash=cash,
                                                                   online=online,
                                                                   shipping=shipping_packing,
                                                                   flat_discount=flat_discount,
                                                                   discount=discount1,
                                                                   discount_amount=discount_amount,
                                                                   pay_amount=after_dis_amount,
                                                                   current=after_dis_amount,
                                                                   state_code=state_code,
                                                                   )
        if obj:
            PatientMedicineBillDetail.objects.filter(head_id=id).exclude(medicine_id__in=medicine_ids).delete()
            status = 'success'
            msg = 'Bill creation Successfully.'

        context = {
            'status': status,
            'msg': msg,
        }
        return JsonResponse(context)

    else:
        user_id = request.session['user_id']
        try:
            store = Store.objects.get(user_id=user_id)
            store_id = store.id
        except:
            store_id = 0

        user = PatientMedicineBillHead.objects.get(id=id)
        is_last = user.id == PatientMedicineBillHead.objects.filter(patient_id=user.patient.id).aggregate(Max('id'))[
            'id__max']
        is_estimated = PatientEstimateMedicineBillHead.objects.filter(head_id=id, patient_id=user.patient.id).exists()
        if is_last and is_estimated == False:
            is_last = True
        else:
            is_last = False
        cash_online_amount = \
            PatientMedicineBillHead.objects.filter(patient_id=user.patient.id).exclude(id=id).aggregate(
                total=Sum(
                    Coalesce(F('cash'), 0) +
                    Coalesce(F('online'), 0) +
                    Coalesce(F('extra_cash_amount'), 0) +
                    Coalesce(F('extra_online_amount'), 0),
                    output_field=DecimalField()
                )
            )['total'] or 0
        total_pay_amount = PatientMedicineBillHead.objects.filter(patient_id=user.patient.id).exclude(id=id).aggregate(
            total=Sum('pay_amount')
        )['total'] or 0

        old_credit_sum = total_pay_amount - cash_online_amount

        medicine = PatientMedicineBillDetail.objects.filter(head_id=id)
        medicine_list = []

        for i in medicine:
            data_dict = {}
            query = Q(to_store_id=store_id, medicine_id=i.medicine.id)
            store_medicine = MedicineStore.objects.filter(query).values('qty', 'price', 'expiry', 'batch_no')
            data_dict['medicine_id'] = i.medicine.id
            data_dict['medicine_name'] = i.medicine.name
            try:
                data_dict['batch_no'] = store_medicine[0]['batch_no']
            except:
                data_dict['batch_no'] = ''
            data_dict['record_qty'] = i.record_qty
            data_dict['sell_qty'] = i.sell_qty
            data_dict['order_qty'] = i.order_qty
            data_dict['discount'] = i.discount
            data_dict['hsn'] = i.hsn
            data_dict['gst'] = i.gst
            data_dict['taxable_amount'] = i.taxable_amount
            data_dict['tax'] = i.tax
            data_dict['expiry'] = store_medicine[0]['expiry']

            try:
                data_dict['record_qty'] = store_medicine[0]['qty']
            except:
                data_dict['record_qty'] = 0
            try:
                data_dict['mrp'] = store_medicine[0]['price']
            except:
                data_dict['mrp'] = 0

            medicine_list.append(data_dict)

        context = {
            'id': id,
            'order_type': order_type,
            'store_id': store_id,
            'user': user,
            'medicine': medicine_list,
            'old_credit_sum': old_credit_sum,
            'is_last': is_last,
        }

        if order_type == 1:
            return render(request, 'patient_update_bill/update_medicine_order_bill_instate.html', context)
        elif order_type == 2:
            return render(request, 'patient_update_bill/update_medicine_order_bill_other_state.html', context)
        elif order_type == 3:
            return render(request, 'patient_update_bill/update_order_bill_of_supply.html', context)
        else:
            return render(request, 'patient_update_bill/update_order_bill_of_supply.html', context)


@login_required(login_url='/account/user_login/')
def estimate_medicine_order_bill(request, order_type, id):
    if request.method == 'POST':
        form = request.POST
        user_id = request.session['user_id']
        try:
            store = Store.objects.get(user_id=user_id)
            store_id = store.id
        except:
            store_id = 0

        status = 'failed'
        msg = 'Bill Creation failed.'

        medicines = json.loads(request.POST.get('medicines'))
        patient_id = form.get('patient_id')
        oder_id = form.get('oder_id')
        invoice_number = estimate_generate_invoice_number()

        flat_discount = float(form.get('flat_discount'))
        subtotal = float(form.get('sub_total'))
        discount = int(form.get('total_discount'))
        shipping_packing = float(form.get('shipping_packing'))
        current = float(form.get('current'))

        discount_amount = subtotal * discount / 100
        after_dis_amount = subtotal - discount_amount + shipping_packing
        after_dis_amount = after_dis_amount - current
        after_dis_amount = after_dis_amount - flat_discount

        cash = float(form.get('cash'))
        online = float(form.get('online'))

        sgst = form.get('sgst')
        cgst = form.get('cgst')

        obj = PatientEstimateMedicineBillHead.objects.create(head_id=id,
                                                             patient_id=patient_id,
                                                             store_id=store_id,
                                                             sgst=sgst,
                                                             cgst=cgst,
                                                             subtotal=subtotal,
                                                             current=current,
                                                             cash=cash,
                                                             online=online,
                                                             shipping=shipping_packing,
                                                             flat_discount=flat_discount,
                                                             discount=discount,
                                                             discount_amount=discount_amount,
                                                             pay_amount=after_dis_amount,
                                                             status=1,
                                                             order_type=order_type,
                                                             )
        if obj:
            head_id = obj.id
            for medicine_data in medicines:
                medicine_id = medicine_data['medicine_id']
                batch_no = medicine_data['batch_no']
                record_qty = int(medicine_data['record_qty'])
                sell_qty = int(medicine_data['sell_qty'])
                discount = int(medicine_data['discount'])
                mrp = float(medicine_data['mrp'])
                sale_rate = float(medicine_data['sale_rate'])
                try:
                    hsn = medicine_data['hsn']
                except:
                    hsn = 0

                try:
                    gst = int(medicine_data['gst'])
                except:
                    gst = 0

                try:
                    taxable_amount = float(medicine_data['taxable_amount'])
                except:
                    taxable_amount = 0

                try:
                    tax = float(medicine_data['tax'])
                except:
                    tax = 0
                amount = float(medicine_data['amount'])

                PatientEstimateMedicineBillDetail.objects.create(head_id=head_id,
                                                                 medicine_id=medicine_id,
                                                                 record_qty=record_qty,
                                                                 sell_qty=sell_qty,
                                                                 mrp=mrp,
                                                                 discount=discount,
                                                                 sale_rate=sale_rate,
                                                                 hsn=hsn,
                                                                 gst=gst,
                                                                 taxable_amount=taxable_amount,
                                                                 tax=tax,
                                                                 amount=amount,
                                                                 batch_no=batch_no,
                                                                 )
                if sell_qty < record_qty:
                    remaining_qty = int(record_qty) - int(sell_qty)
                    MedicineStore.objects.filter(to_store_id=store_id, medicine_id=medicine_id,
                                                 batch_no=batch_no).update(
                        qty=remaining_qty)
            PatientMedicineBillHead.objects.filter(id=id).update(status=1, estimate_status=1)
            PatientEstimateMedicineBillHead.filter(id=head_id).update(invoice_number=invoice_number)
            status = 'success'
            msg = 'Bill creation Successfully.'

        context = {
            'status': status,
            'msg': msg,
        }
        return JsonResponse(context)

    else:
        user_id = request.session['user_id']
        try:
            store = Store.objects.get(user_id=user_id)
            store_id = store.id
        except:
            store_id = 0

        user = PatientMedicineBillHead.objects.get(id=id)
        invoice_number = user.invoice_number

        cash_online_amount = PatientEstimateMedicineBillHead.objects.filter(patient_id=user.patient.id).aggregate(
            total=Sum(
                Coalesce(F('cash'), 0) +
                Coalesce(F('online'), 0) +
                Coalesce(F('extra_cash_amount'), 0) +
                Coalesce(F('extra_online_amount'), 0),
                output_field=DecimalField()
            )
        )['total'] or 0

        total_pay_amount = PatientEstimateMedicineBillHead.objects.filter(patient_id=user.patient.id).aggregate(
            total=Sum('pay_amount')
        )['total'] or 0

        old_credit_sum = total_pay_amount - cash_online_amount

        medicine = PatientMedicineBillDetail.objects.filter(head_id=id)
        medicine_list = []
        for i in medicine:
            data_dict = {}
            query = Q(to_store_id=store_id, medicine_id=i.medicine.id)
            store_medicine = MedicineStore.objects.filter(query).values('qty', 'price', 'expiry', 'batch_no')
            data_dict['medicine_id'] = i.medicine.id
            data_dict['medicine_name'] = i.medicine.name
            try:
                data_dict['batch_no'] = store_medicine[0]['batch_no']
            except:
                data_dict['batch_no'] = ''

            data_dict['order_qty'] = i.order_qty
            data_dict['sell_qty'] = i.sell_qty
            data_dict['discount'] = i.discount
            data_dict['hsn'] = i.hsn
            data_dict['gst'] = i.gst
            data_dict['taxable_amount'] = i.taxable_amount
            data_dict['tax'] = i.tax
            try:
                data_dict['expiry'] = store_medicine[0]['expiry']
            except:
                data_dict['expiry'] = ''

            try:
                data_dict['record_qty'] = store_medicine[0]['qty']
            except:
                data_dict['record_qty'] = 0
            try:
                data_dict['mrp'] = store_medicine[0]['price']
            except:
                data_dict['mrp'] = 0

            medicine_list.append(data_dict)
        context = {
            'id': id,
            'order_type': order_type,
            'store_id': store_id,
            'user': user,
            'medicine': medicine_list,
            'old_credit_sum': old_credit_sum,
            'invoice_number': invoice_number,
        }
        if order_type == 1:
            return render(request, 'patient_estimate/estimate_medicine_order_bill_instate.html', context)
        elif order_type == 2:
            return render(request, 'patient_estimate/estimate_medicine_order_bill_other_state.html', context)
        elif order_type == 3:
            return render(request, 'patient_estimate/estimate_medicine_order_bill_of_supply.html', context)
        else:
            return render(request, 'patient_estimate/estimate_medicine_order_bill_of_supply.html', context)


@login_required(login_url='/account/user_login/')
def update_estimate_medicine_order_bill(request, order_type, id):
    if request.method == 'POST':
        form = request.POST
        user_id = request.session['user_id']
        try:
            store = Store.objects.get(user_id=user_id)
            store_id = store.id
        except:
            store_id = 0

        status = 'failed'
        msg = 'Estimated Bill updated failed.'
        medicines = json.loads(request.POST.get('medicines'))

        subtotal = float(form.get('sub_total'))
        flat_discount = float(form.get('flat_discount'))
        discount1 = int(form.get('total_discount'))
        shipping_packing = float(form.get('shipping_packing'))
        current = float(form.get('current'))

        discount_amount = subtotal * discount1 / 100
        after_dis_amount = subtotal - discount_amount + shipping_packing
        after_dis_amount = after_dis_amount - current
        after_dis_amount = after_dis_amount - flat_discount

        cash = float(form.get('cash'))
        online = float(form.get('online'))

        sgst = form.get('sgst')
        cgst = form.get('cgst')

        head_id = id
        medicine_ids = []
        for medicine_data in medicines:
            medicine_id = medicine_data['medicine_id']
            medicine_ids.append(medicine_id)
            record_qty = int(medicine_data['record_qty'])
            sell_qty = int(medicine_data['sell_qty'])
            discount = int(medicine_data['discount'])
            mrp = float(medicine_data['mrp'])
            sale_rate = float(medicine_data['sale_rate'])
            try:
                hsn = medicine_data['hsn']
            except:
                hsn = 0

            try:
                gst = int(medicine_data['gst'])
            except:
                gst = 0

            try:
                taxable_amount = float(medicine_data['taxable_amount'])
            except:
                taxable_amount = 0

            try:
                tax = float(medicine_data['tax'])
            except:
                tax = 0
            amount = float(medicine_data['amount'])
            query = Q(head_id=head_id, medicine_id=medicine_id)
            already_obj = PatientEstimateMedicineBillDetail.objects.filter(query)
            if already_obj:
                PatientEstimateMedicineBillDetail.objects.filter(query).update(record_qty=record_qty,
                                                                               sell_qty=sell_qty,
                                                                               mrp=mrp,
                                                                               discount=discount,
                                                                               sale_rate=sale_rate,
                                                                               hsn=hsn,
                                                                               gst=gst,
                                                                               taxable_amount=taxable_amount,
                                                                               tax=tax,
                                                                               amount=amount,
                                                                               )
        obj = PatientEstimateMedicineBillHead.objects.filter(id=id).update(store_id=store_id,
                                                                           sgst=sgst,
                                                                           cgst=cgst,
                                                                           subtotal=subtotal,
                                                                           current=current,
                                                                           cash=cash,
                                                                           online=online,
                                                                           shipping=shipping_packing,
                                                                           flat_discount=flat_discount,
                                                                           discount=discount1,
                                                                           discount_amount=discount_amount,
                                                                           pay_amount=after_dis_amount,
                                                                           )
        if obj:
            PatientEstimateMedicineBillHead.objects.filter(id=id).update(status=1)
            PatientEstimateMedicineBillDetail.objects.filter(head_id=id).exclude(medicine_id__in=medicine_ids).delete()
            status = 'success'
            msg = 'Estimated Bill updated Successfully.'

        context = {
            'status': status,
            'msg': msg,
        }
        return JsonResponse(context)

    else:
        user_id = request.session['user_id']
        try:
            store = Store.objects.get(user_id=user_id)
            store_id = store.id
        except:
            store_id = 0

        user = PatientEstimateMedicineBillHead.objects.get(id=id)

        is_last = user.id == \
                  PatientEstimateMedicineBillHead.objects.filter(patient_id=user.patient.id).aggregate(Max('id'))[
                      'id__max']
        if is_last:
            is_last = True
        else:
            is_last = False

        cash_online_amount = \
            PatientEstimateMedicineBillHead.objects.filter(patient_id=user.patient.id).exclude(id=id).aggregate(
                total=Sum(
                    Coalesce(F('cash'), 0) +
                    Coalesce(F('online'), 0) +
                    Coalesce(F('extra_cash_amount'), 0) +
                    Coalesce(F('extra_online_amount'), 0),
                    output_field=DecimalField()
                )
            )['total'] or 0

        total_pay_amount = \
            PatientEstimateMedicineBillHead.objects.filter(patient_id=user.patient.id).exclude(id=id).aggregate(
                total=Sum('pay_amount')
            )['total'] or 0

        old_credit_sum = total_pay_amount - cash_online_amount

        medicine = PatientEstimateMedicineBillDetail.objects.filter(head_id=id)
        medicine_count = PatientEstimateMedicineBillDetail.objects.filter(head_id=id).count()
        medicine_list = []

        for i in medicine:
            data_dict = {}
            query = Q(to_store_id=store_id, medicine_id=i.medicine.id)
            store_medicine = MedicineStore.objects.filter(query).values('qty', 'price', 'expiry', 'batch_no')
            data_dict['medicine_id'] = i.medicine.id
            data_dict['medicine_name'] = i.medicine.name
            try:
                data_dict['batch_no'] = store_medicine[0]['batch_no']
            except:
                data_dict['batch_no'] = ''
            data_dict['order_qty'] = i.order_qty
            data_dict['sell_qty'] = i.sell_qty
            data_dict['discount'] = i.discount
            data_dict['hsn'] = i.hsn
            data_dict['gst'] = i.gst
            data_dict['taxable_amount'] = i.taxable_amount
            data_dict['tax'] = i.tax
            try:
                data_dict['expiry'] = store_medicine[0]['expiry']
            except:
                data_dict['expiry'] = ''

            try:
                data_dict['record_qty'] = store_medicine[0]['qty']
            except:
                data_dict['record_qty'] = 0
            try:
                data_dict['mrp'] = store_medicine[0]['price']
            except:
                data_dict['mrp'] = 0

            medicine_list.append(data_dict)

        context = {
            'id': id,
            'order_type': order_type,
            'store_id': store_id,
            'user': user,
            'medicine': medicine_list,
            'old_credit_sum': old_credit_sum,
            'is_last': is_last,
        }
        if order_type == 1:
            return render(request, 'patient_update_estimate/update_estimate_medicine_order_bill_instate.html', context)
        elif order_type == 2:
            return render(request, 'patient_update_estimate/update_estimate_medicine_order_bill_other_state.html',
                          context)
        elif order_type == 3:
            return render(request, 'patient_update_estimate/update_estimate_medicine_order_bill_of_supply.html',
                          context)
        else:
            return render(request, 'patient_update_estimate/update_estimate_medicine_order_bill_of_supply.html',
                          context)


def normal_order_bill_list(request):
    order = PatientMedicineBillHead.objects.all().order_by('-id')
    context = {
        'order': order
    }
    return render(request, 'patient_bill_list/normal_order_bill_list.html', context)


def estimate_order_bill_list(request):
    order = PatientEstimateMedicineBillHead.objects.filter(status=1).order_by('-id')
    context = {
        'order': order
    }
    return render(request, 'patient_bill_list/estimate_order_bill_list.html', context)


def view_normal_invoice(request, id):
    user_id = request.session['user_id']
    try:
        store = Store.objects.get(user_id=user_id)
        store_id = store.id
    except:
        store_id = 0
    try:
        user = PatientMedicineBillHead.objects.get(id=id)
    except:
        return redirect('/my_order/my_medicine_ordered_list/')

    order_type = user.order_type
    invoice_number = user.invoice_number
    cash_online_amount = PatientMedicineBillHead.objects.filter(patient_id=user.patient.id).exclude(id=id).aggregate(
        total=Sum(
            Coalesce(F('cash'), 0) +
            Coalesce(F('online'), 0) +
            Coalesce(F('extra_cash_amount'), 0) +
            Coalesce(F('extra_online_amount'), 0),
            output_field=DecimalField()
        )
    )['total'] or 0

    total_pay_amount = PatientMedicineBillHead.objects.filter(patient_id=user.patient.id).exclude(id=id).aggregate(
        total=Sum('pay_amount')
    )['total'] or 0

    old_credit_sum = total_pay_amount - cash_online_amount

    pay_amount = user.pay_amount
    cash = user.cash
    online = user.online

    remaining_amount = pay_amount - (cash + online) + old_credit_sum
    current = user.current

    medicine = PatientMedicineBillDetail.objects.filter(head_id=user.id)
    gst_per = PatientMedicineBillDetail.objects.filter(head_id=user.id).values_list('gst', flat=True).distinct()

    total_taxable_by_gst_list = []
    grand_sub_total = 0
    grand_discount_total = 0
    grand_taxable_amount_total = 0
    grand_sgst_and_cgst_total = 0
    grand_tax_total = 0
    for gst in gst_per:
        # Total taxable amounts
        total_taxable = PatientMedicineBillDetail.objects.filter(
            head_id=user.id, gst=gst
        ).aggregate(total_taxable=Sum('taxable_amount'))['total_taxable'] or 0
        grand_taxable_amount_total += total_taxable

        # Sub totals
        medicine_sub_totals = PatientMedicineBillDetail.objects.filter(
            head_id=user.id, gst=gst
        ).annotate(
            total=F('sell_qty') * F('mrp'),
            discount_amount=ExpressionWrapper(
                (F('sell_qty') * F('mrp') * F('discount') / 100),
                output_field=FloatField()
            )
        )

        # Sub total calculation
        sub_total = medicine_sub_totals.aggregate(grand_total=Sum('total'))['grand_total'] or 0
        grand_sub_total += sub_total

        # Discount total calculation
        discount_total = medicine_sub_totals.aggregate(grand_discount=Sum('discount_amount'))['grand_discount'] or 0
        grand_discount_total += discount_total

        tax = PatientMedicineBillDetail.objects.filter(
            head_id=user.id, gst=gst
        ).aggregate(tax=Sum('tax'))['tax'] or 0
        grand_tax_total += tax

        if order_type == 1:
            sgst_and_cgst = tax / 2
        elif order_type == 2:
            sgst_and_cgst = tax
        else:
            sgst_and_cgst = tax

        grand_sgst_and_cgst_total += sgst_and_cgst
        total_taxable_by_gst_list.append({
            'gst': gst,
            'taxable_amount': total_taxable,
            'sub_total': sub_total,
            'discount': discount_total,
            'tax': tax,
            'sgst_and_cgst': sgst_and_cgst,
        })

    total_discounted_price = 0
    for item in medicine:
        if item.mrp and item.discount and item.sell_qty:  # Ensure values are not None
            discount_amount = (item.mrp * item.discount) / 100  # Calculate discount amount per unit
            discounted_price_per_unit = discount_amount  # Calculate discounted price per unit
            total_price_for_item = discounted_price_per_unit * item.sell_qty  # Multiply by sell quantity
            total_discounted_price += total_price_for_item  # Add to total discounted price

    subtotal_amount = medicine.aggregate(total_amount=Sum(F('sell_qty') * F('mrp')))['total_amount'] or 0
    total_taxable_amount = medicine.aggregate(total_taxable=Sum('taxable_amount'))['total_taxable'] or 0

    context = {
        'id': id,
        'order_type': order_type,
        'store_id': store_id,
        'user': user,
        'medicine': medicine,
        'total_taxable_by_gst': total_taxable_by_gst_list,
        'total_discounted_price': total_discounted_price,
        'subtotal_amount': subtotal_amount,
        'total_taxable_amount': total_taxable_amount,

        'grand_sub_total': grand_sub_total,
        'grand_discount_total': grand_discount_total,
        'grand_taxable_amount_total': grand_taxable_amount_total,
        'grand_sgst_and_cgst_total': grand_sgst_and_cgst_total,
        'grand_tax_total': grand_tax_total,

        'remaining_amount': remaining_amount,
        'current': current,
        'old_credit_sum': old_credit_sum,
        'invoice_number': invoice_number,
    }

    if order_type == 1:
        return render(request, 'patient_normal_invoice/normal_invoice_instate.html', context)
    elif order_type == 2:
        return render(request, 'patient_normal_invoice/normal_invoice_other_state.html', context)
    elif order_type == 3:
        return render(request, 'patient_normal_invoice/normal_invoice_bill_of_supply.html', context)
    else:
        return render(request, 'patient_normal_invoice/normal_invoice_bill_of_supply.html', context)


def final_bill_invoice(request, id):
    user_id = request.session['user_id']
    try:
        store = Store.objects.get(user_id=user_id)
        store_id = store.id
    except:
        store_id = 0
    try:
        user = PatientMedicineUnregisteredBillHead.objects.get(head_id=id)
    except:
        return redirect('/patient/normal_order_bill_list/')

    order_type = user.order_type
    invoice_number = user.invoice_number
    cash_online_amount = \
        PatientMedicineUnregisteredBillHead.objects.filter(patient_id=user.patient.id).exclude(head_id=id).aggregate(
            total=Sum(
                Coalesce(F('cash'), 0) +
                Coalesce(F('online'), 0) +
                Coalesce(F('extra_cash_amount'), 0) +
                Coalesce(F('extra_online_amount'), 0),
                output_field=DecimalField()
            )
        )['total'] or 0

    total_pay_amount = \
        PatientMedicineUnregisteredBillHead.objects.filter(patient_id=user.patient.id).exclude(head_id=id).aggregate(
            total=Sum('pay_amount')
        )['total'] or 0

    old_credit_sum = total_pay_amount - cash_online_amount

    pay_amount = user.pay_amount
    cash = user.cash
    online = user.online

    remaining_amount = pay_amount - (cash + online) + old_credit_sum
    current = user.current

    medicine = PatientMedicineUnregisteredBillDetail.objects.filter(head_id=user.id)
    gst_per = PatientMedicineUnregisteredBillDetail.objects.filter(head_id=user.id).values_list('gst',
                                                                                                flat=True).distinct()

    total_taxable_by_gst_list = []
    grand_sub_total = 0
    grand_discount_total = 0
    grand_taxable_amount_total = 0
    grand_sgst_and_cgst_total = 0
    grand_tax_total = 0
    for gst in gst_per:
        # Total taxable amounts
        total_taxable = PatientMedicineUnregisteredBillDetail.objects.filter(
            head_id=user.id, gst=gst
        ).aggregate(total_taxable=Sum('taxable_amount'))['total_taxable'] or 0
        grand_taxable_amount_total += total_taxable

        # Sub totals
        medicine_sub_totals = PatientMedicineUnregisteredBillDetail.objects.filter(
            head_id=user.id, gst=gst
        ).annotate(
            total=F('sell_qty') * F('mrp'),
            discount_amount=ExpressionWrapper(
                (F('sell_qty') * F('mrp') * F('discount') / 100),
                output_field=FloatField()
            )
        )

        # Sub total calculation
        sub_total = medicine_sub_totals.aggregate(grand_total=Sum('total'))['grand_total'] or 0
        grand_sub_total += sub_total

        # Discount total calculation
        discount_total = medicine_sub_totals.aggregate(grand_discount=Sum('discount_amount'))['grand_discount'] or 0
        grand_discount_total += discount_total

        tax = PatientMedicineUnregisteredBillDetail.objects.filter(
            head_id=user.id, gst=gst
        ).aggregate(tax=Sum('tax'))['tax'] or 0
        grand_tax_total += tax

        if order_type == 1:
            sgst_and_cgst = tax / 2
        elif order_type == 2:
            sgst_and_cgst = tax
        else:
            sgst_and_cgst = tax

        grand_sgst_and_cgst_total += sgst_and_cgst
        total_taxable_by_gst_list.append({
            'gst': gst,
            'taxable_amount': total_taxable,
            'sub_total': sub_total,
            'discount': discount_total,
            'tax': tax,
            'sgst_and_cgst': sgst_and_cgst,
        })

    total_discounted_price = 0
    for item in medicine:
        if item.mrp and item.discount and item.sell_qty:  # Ensure values are not None
            discount_amount = (item.mrp * item.discount) / 100  # Calculate discount amount per unit
            discounted_price_per_unit = discount_amount  # Calculate discounted price per unit
            total_price_for_item = discounted_price_per_unit * item.sell_qty  # Multiply by sell quantity
            total_discounted_price += total_price_for_item  # Add to total discounted price

    subtotal_amount = medicine.aggregate(total_amount=Sum(F('sell_qty') * F('mrp')))['total_amount'] or 0
    total_taxable_amount = medicine.aggregate(total_taxable=Sum('taxable_amount'))['total_taxable'] or 0

    context = {
        'id': id,
        'order_type': order_type,
        'store_id': store_id,
        'user': user,
        'medicine': medicine,
        'total_taxable_by_gst': total_taxable_by_gst_list,
        'total_discounted_price': total_discounted_price,
        'subtotal_amount': subtotal_amount,
        'total_taxable_amount': total_taxable_amount,

        'grand_sub_total': grand_sub_total,
        'grand_discount_total': grand_discount_total,
        'grand_taxable_amount_total': grand_taxable_amount_total,
        'grand_sgst_and_cgst_total': grand_sgst_and_cgst_total,
        'grand_tax_total': grand_tax_total,

        'remaining_amount': remaining_amount,
        'current': current,
        'old_credit_sum': old_credit_sum,
        'invoice_number': invoice_number,
    }

    if order_type == 1:
        return render(request, 'patient_final_invoice/normal_invoice_instate.html', context)
    elif order_type == 2:
        return render(request, 'patient_final_invoice/normal_invoice_other_state.html', context)
    elif order_type == 3:
        return render(request, 'patient_final_invoice/normal_invoice_bill_of_supply.html', context)
    else:
        return render(request, 'patient_final_invoice/normal_invoice_bill_of_supply.html', context)


def view_estimate_invoice(request, id):
    user_id = request.session['user_id']
    try:
        store = Store.objects.get(user_id=user_id)
        store_id = store.id
    except:
        store_id = 0
    try:
        user = PatientEstimateMedicineBillHead.objects.get(id=id)
    except:
        return redirect('/patient/estimate_order_bill_list/')
    order_type = user.order_type

    cash_online_amount = \
        PatientEstimateMedicineBillHead.objects.filter(patient_id=user.patient.id).exclude(id=id).aggregate(
            total=Sum(
                Coalesce(F('cash'), 0) +
                Coalesce(F('online'), 0) +
                Coalesce(F('extra_cash_amount'), 0) +
                Coalesce(F('extra_online_amount'), 0),
                output_field=DecimalField()
            )
        )['total'] or 0

    total_pay_amount = \
        PatientEstimateMedicineBillHead.objects.filter(patient_id=user.patient.id).exclude(id=id).aggregate(
            total=Sum('pay_amount')
        )['total'] or 0

    old_credit_sum = total_pay_amount - cash_online_amount
    pay_amount = user.pay_amount
    cash = user.cash
    online = user.online
    total_amt = user.subtotal - user.discount_amount

    remaining_amount = pay_amount - (cash + online) + old_credit_sum

    current = user.current

    medicine = PatientEstimateMedicineBillDetail.objects.filter(head_id=user.id)

    gst_per = PatientEstimateMedicineBillDetail.objects.filter(head_id=user.id).values_list('gst', flat=True).distinct()

    total_taxable_by_gst_list = []
    grand_sub_total = 0
    grand_discount_total = 0
    grand_taxable_amount_total = 0
    grand_sgst_and_cgst_total = 0
    grand_tax_total = 0

    for gst in gst_per:
        # Total taxable amounts
        total_taxable = PatientEstimateMedicineBillDetail.objects.filter(
            head_id=user.id, gst=gst
        ).aggregate(total_taxable=Sum('taxable_amount'))['total_taxable'] or 0
        grand_taxable_amount_total += total_taxable

        # Sub totals
        medicine_sub_totals = PatientEstimateMedicineBillDetail.objects.filter(
            head_id=user.id, gst=gst
        ).annotate(
            total=F('sell_qty') * F('mrp'),
            discount_amount=ExpressionWrapper(
                (F('sell_qty') * F('mrp') * F('discount') / 100),
                output_field=FloatField()
            )
        )

        # Sub total calculation
        sub_total = medicine_sub_totals.aggregate(grand_total=Sum('total'))['grand_total'] or 0
        grand_sub_total += sub_total

        # Discount total calculation
        discount_total = medicine_sub_totals.aggregate(grand_discount=Sum('discount_amount'))['grand_discount'] or 0
        grand_discount_total += discount_total

        tax = PatientEstimateMedicineBillDetail.objects.filter(
            head_id=user.id, gst=gst
        ).aggregate(tax=Sum('tax'))['tax'] or 0
        grand_tax_total += tax
        sgst_and_cgst = tax / 2
        grand_sgst_and_cgst_total += sgst_and_cgst
        total_taxable_by_gst_list.append({
            'gst': gst,
            'taxable_amount': total_taxable,
            'sub_total': sub_total,
            'discount': discount_total,
            'tax': tax,
            'sgst_and_cgst': sgst_and_cgst,
        })

    total_discounted_price = 0
    for item in medicine:
        if item.mrp and item.discount and item.sell_qty:  # Ensure values are not None
            discount_amount = (item.mrp * item.discount) / 100  # Calculate discount amount per unit
            discounted_price_per_unit = discount_amount  # Calculate discounted price per unit
            total_price_for_item = discounted_price_per_unit * item.sell_qty  # Multiply by sell quantity
            total_discounted_price += total_price_for_item  # Add to total discounted price

    subtotal_amount = medicine.aggregate(total_amount=Sum(F('sell_qty') * F('mrp')))['total_amount'] or 0
    total_taxable_amount = medicine.aggregate(total_taxable=Sum('taxable_amount'))['total_taxable'] or 0

    context = {
        'id': id,
        'order_type': order_type,
        'store_id': store_id,
        'user': user,
        'medicine': medicine,
        'old_credit_sum': old_credit_sum,
        'total_taxable_by_gst': total_taxable_by_gst_list,
        'total_discounted_price': total_discounted_price,
        'subtotal_amount': subtotal_amount,
        'total_taxable_amount': total_taxable_amount,

        'grand_sub_total': grand_sub_total,
        'grand_discount_total': grand_discount_total,
        'grand_taxable_amount_total': grand_taxable_amount_total,
        'grand_sgst_and_cgst_total': grand_sgst_and_cgst_total,
        'grand_tax_total': grand_tax_total,

        'current': current,
        'pay_amount': pay_amount,
        'remaining_amount': remaining_amount,
        'total_amt': total_amt,
    }
    if order_type == 1:
        return render(request, 'patient_estimate_invoice/estimate_invoice_instate.html', context)
    elif order_type == 2:
        return render(request, 'patient_estimate_invoice/estimate_invoice_other_state.html', context)
    elif order_type == 3:
        return render(request, 'patient_estimate_invoice/estimate_invoice_bill_of_supply.html', context)
    else:
        return render(request, 'patient_estimate_invoice/estimate_invoice_bill_of_supply.html', context)


def add_extra_amount(request, id):
    if request.method == 'POST':
        form = request.POST
        cash_amount = form.get('cash_amount')
        online_amount = form.get('online_amount')
        amount_remark = form.get('amount_remark')

        if cash_amount:
            cash_amount = Decimal(cash_amount)
        else:
            cash_amount = 0

        if online_amount:
            online_amount = Decimal(online_amount)
        else:
            online_amount = 0

        record = PatientMedicineBillHead.objects.filter(id=id)
        if record:
            PatientMedicineBillHead.objects.filter(id=id).update(extra_cash_amount=cash_amount,
                                                                 extra_online_amount=online_amount,
                                                                 )

            status = 'success'

            context = {
                'status': status,
                'msg': 'Extra amount added successfully.',
            }
            return JsonResponse(context)


def estimate_add_extra_amount(request, id):
    if request.method == 'POST':
        form = request.POST
        cash_amount = form.get('cash_amount')
        online_amount = form.get('online_amount')
        amount_remark = form.get('amount_remark')

        if cash_amount:
            cash_amount = Decimal(cash_amount)
        else:
            cash_amount = 0

        if online_amount:
            online_amount = Decimal(online_amount)
        else:
            online_amount = 0

        record = PatientEstimateMedicineBillHead.objects.filter(id=id)
        if record:
            PatientEstimateMedicineBillHead.objects.filter(id=id).update(extra_cash_amount=cash_amount,
                                                                         extra_online_amount=online_amount,
                                                                         )

            status = 'success'

            context = {
                'status': status,
                'msg': 'Extra amount added successfully.',
            }
            return JsonResponse(context)


def export_today_patient_bills_excel(request):
    user_id = request.session['user_id']
    try:
        store = Store.objects.get(user_id=user_id)
        store_id = store.id
    except:
        store_id = 0
    today = timezone.now().date()
    start_datetime = timezone.make_aware(datetime.combine(today, time.min))
    end_datetime = timezone.make_aware(datetime.combine(today, time.max))

    bills = PatientMedicineBillHead.objects.filter(store_id=store_id, created_at__range=(start_datetime, end_datetime))

    wb = Workbook()
    ws = wb.active
    ws.title = "Today's Bills"

    headers = [
        'Date', 'Invoice', 'Name', 'City', 'Subtotal', 'Discount %',
        'Flat', 'Pay Amount', 'Old Credit', 'New Credit', 'Cash', 'Online',
        'Extra Cash Amount', 'Extra Online Amount', 'Store', 'Patient',
    ]
    ws.append(headers)

    # Initialize total trackers
    total_fields = {
        'subtotal': 0,
        'discount_amount': 0,
        'flat_discount': 0,
        'pay_amount': 0,
        'old_credit': 0,
        'new_credit': 0,
        'cash': 0,
        'online': 0,
        'extra_cash_amount': 0,
        'extra_online_amount': 0,
    }

    for bill in bills:
        row = [
            bill.created_at.strftime('%d-%m-%Y %I:%M:%S %p') if bill.created_at else '',
            bill.invoice_number,
            bill.patient.user.first_name,
            bill.patient.user.city,
            bill.subtotal,
            bill.discount_amount,
            bill.flat_discount,
            bill.pay_amount,
            bill.old_credit,
            bill.new_credit,
            bill.cash,
            bill.online,
            bill.extra_cash_amount,
            bill.extra_online_amount,
            str(bill.store) if bill.store else '',
            str(bill.patient) if bill.patient else '',

        ]
        ws.append(row)

        # Update totals
        for key in total_fields.keys():
            value = getattr(bill, key, 0) or 0
            total_fields[key] += float(value)

    # Append total row
    total_row = ['Grand Total', '']
    # Append totals in the correct position
    for key in [
        'subtotal', 'discount_amount', 'flat_discount', 'pay_amount',
        'old_credit', 'new_credit', 'cash', 'online', 'extra_cash_amount', 'extra_online_amount'
    ]:
        if key in total_fields:
            total_row.append(round(total_fields[key], 2))
        elif key == 'discount':  # Discount is an IntegerField, not summed above
            total_row.append(sum([bill.discount or 0 for bill in bills]))
        else:
            total_row.append('')

    # Fill remaining cells for alignment
    while len(total_row) < len(headers):
        total_row.append('')

    ws.append([])
    ws.append(total_row)

    # # ✅ Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Today_Patient_Bills_{today.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response
